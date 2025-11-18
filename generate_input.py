import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import _lib.utils

# import warnings
# warnings.filterwarnings("ignore", category=FutureWarning)


serverDb = _lib.utils.serverDb


def generate_input(driver):
	
	months = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']
	# detection = ['DETECTIONRATE','DETECTIONTREATMENTRATE','PREVENTATIVETREATMENTRATE']
	detection = ['SYSTEMMITIGATIONRATE','PREVENTATIVETREATMENTRATE','DETECTIONRATE','DETECTIONTREATMENTRATE']

	suitability = ['FAVOURABLE','SUITABLE','MARGINAL','UNSUITABLE']
	# consequences = ['ECONESTAB','ECONSPREAD','ENVIRONESTAB','ENVIRONSPREAD','HEALTHESTAB','HEALTHSPREAD','SOCIALESTAB','SOCIALSPREAD']
	consequences = ['ECON','ENVIRON','HEALTH','SOCIAL']


	with serverDb() as db:
		pathwayPoint_df = pd.read_sql_query("SELECT pathway as PATHWAY, name as PATHWAYPOINT FROM pathwayPoint", db.conn)
		landCover_df = pd.read_sql_query("SELECT name as LANDCOVER FROM landCover", db.conn)


	df = pd.read_excel(driver, sheet_name='context')
	df = df.merge(pathwayPoint_df, how='left', on='PATHWAY')

	wb = Workbook()
	wb.remove(wb.active)
	
	
	dataframes = {
		'units': pd.DataFrame(columns=['PATHWAY', 'SOURCE', *months]),
		'itemsPerUnit': pd.DataFrame(columns=['PATHWAY', 'SOURCE', 'ITEM', *months]),
		'hostsPerItem': pd.DataFrame(columns=['SOURCE', 'HOST', 'ITEM', *months]), # Dropped Pathway
		'pestsPerHost': pd.DataFrame(columns=['SOURCE', 'PEST', 'HOST', *months]), # Dropped Pathway & Host
		
		'preborderDetectionItem': pd.DataFrame(columns=['PATHWAY', 'SOURCE', 'ITEM', *detection]),
		'preborderDetectionHost': pd.DataFrame(columns=['PATHWAY', 'SOURCE', 'HOST', *detection]), # Dropped Item
		'preborderDetectionPest': pd.DataFrame(columns=['PATHWAY', 'SOURCE', 'PEST', *detection]), # Dropped Host & Item
		
		'pathwayDetectionItem': pd.DataFrame(columns=['PATHWAY', 'PATHWAYPOINT', 'ITEM', *detection]),
		'pathwayDetectionHost': pd.DataFrame(columns=['PATHWAY', 'PATHWAYPOINT', 'HOST', *detection]), # Dropped Item
		'pathwayDetectionPest': pd.DataFrame(columns=['PATHWAY', 'PATHWAYPOINT', 'PEST', *detection]), # Dropped Host & Item
		
		'mortalityItem': pd.DataFrame(columns=['PATHWAY', 'PATHWAYPOINT', 'ITEM', *months]),
		'mortalityHost': pd.DataFrame(columns=['PATHWAY', 'PATHWAYPOINT', 'HOST', 'ITEM', *months]),
		'mortalityPest': pd.DataFrame(columns=['PATHWAY', 'PATHWAYPOINT', 'PEST', 'HOST', 'ITEM', *months]),
		
		'escapeItem': pd.DataFrame(columns=['PATHWAY', 'PATHWAYPOINT', 'ITEM', *months]),
		'escapeHost': pd.DataFrame(columns=['PATHWAY', 'PATHWAYPOINT', 'HOST', 'ITEM', *months]),
		'escapePest': pd.DataFrame(columns=['PATHWAY', 'PATHWAYPOINT', 'PEST', 'HOST', 'ITEM', *months]),
		
		'dispersalHost': pd.DataFrame(columns=['HOST', 'SD']),
		'transmissionHost': pd.DataFrame(columns=['HOST', *suitability]),
		
		'environMortality': pd.DataFrame(columns=[*suitability]),
		'environEstablishment': pd.DataFrame(columns=[*suitability]),
		
		'establishmentDetection': pd.DataFrame(columns=[*detection]),
		'establishmentMortality': pd.DataFrame(columns=[*suitability]),
		'establishmentSpread': pd.DataFrame(columns=[*suitability]),
		
		'landSuitability': pd.DataFrame(columns=['LANDCOVER', 'SUITABILITY']),
		'consequences': pd.DataFrame(columns=['LANDCOVER', *consequences]),
		
		'context': df,
	}


	# updates = {
	# 	'units': df[['PATHWAY', 'SOURCE']],
	# 	'itemsPerUnit': df[['ITEM', 'PATHWAY', 'SOURCE']],
	# 	'hostsPerItem': df[['HOST', 'ITEM', 'SOURCE']],
	# 	'pestsPerHost': df[['PEST', 'HOST', 'SOURCE']],
	# 	'preborderDetectionItem': df[['ITEM', 'PATHWAY', 'SOURCE']],
	# 	'preborderDetectionHost': df[['HOST', 'PATHWAY', 'SOURCE']],
	# 	'preborderDetectionPest': df[['PEST', 'PATHWAY', 'SOURCE']],
	# 	'pathwayDetectionItem': df[['ITEM', 'PATHWAY', 'PATHWAYPOINT']],
	# 	'pathwayDetectionHost': df[['HOST', 'PATHWAY', 'PATHWAYPOINT']],
	# 	'pathwayDetectionPest': df[['PEST', 'PATHWAY', 'PATHWAYPOINT']],
	# 	'mortalityItem': df[['ITEM', 'PATHWAY', 'PATHWAYPOINT']],
	# 	'mortalityHost': df[['HOST', 'PATHWAY', 'PATHWAYPOINT']],
	# 	'mortalityPest': df[['PEST', 'PATHWAY', 'PATHWAYPOINT']],
	# 	'escapeItem':  df[['ITEM', 'PATHWAY', 'PATHWAYPOINT']],
	# 	'escapeHost': df[['HOST', 'PATHWAY', 'PATHWAYPOINT']],
	# 	'escapePest': df[['PEST', 'PATHWAY', 'PATHWAYPOINT']],
	# 	'dispersalHost': df[['HOST']],
	# 	'transmissionHost': df[['HOST']],
	# 	'environMortality': pd.DataFrame([{col: None for col in suitability}]),
	# 	'environEstablishment': pd.DataFrame([{col: None for col in suitability}]),
	# 	'establishmentDetection': pd.DataFrame([{col: None for col in detection}]),
	# 	'establishmentMortality': pd.DataFrame([{col: None for col in suitability}]),
	# 	'establishmentSpread': pd.DataFrame([{col: None for col in suitability}]),
	# 	'landSuitability': landCover_df,
	# 	'consequences': landCover_df,
	# }

	# for key, new_df in updates.items():
	# 	dataframes[key] = pd.concat([dataframes[key], new_df.drop_duplicates().reset_index(drop=True)], ignore_index=True)
		

		
	for key in ['units']:
		dataframes[key] = dataframes[key].fillna(1000)
		
	for key in ['itemsPerUnit','hostsPerItem','pestsPerHost','mortalityItem','mortalityHost','mortalityPest','escapeItem','escapeHost','escapePest','transmissionHost','environMortality','environEstablishment','establishmentMortality','establishmentSpread']:
		dataframes[key] = dataframes[key].fillna(0.01)
		
	for key in ['preborderDetectionItem','preborderDetectionHost','preborderDetectionPest','pathwayDetectionItem','pathwayDetectionHost','pathwayDetectionPest','establishmentDetection','consequences']:
		dataframes[key] = dataframes[key].fillna(0)
		
	for key in['dispersalHost','landSuitability']:
		dataframes[key] = dataframes[key].fillna(1)
		
	

	for sheet_name, df in dataframes.items():
		ws = wb.create_sheet(title=sheet_name)
		ws.append(df.columns.tolist())
		for row in df.itertuples(index=False):
			ws.append(list(row))

		for i, col in enumerate(df.columns, 1):
			max_length = max((len(str(s)) if s is not None else 0) for s in [col] + df[col].astype(str).tolist() )
			ws.column_dimensions[get_column_letter(i)].width = max_length + 2

	wb.save("output.xlsx")
	



generate_input('Liberibacter.xlsx')
















# def generate_input(driver):

# 	with serverDb() as db:
# 		pathwayPoint_df = pd.read_sql_query("SELECT item as ITEM, itemId as ITEMID, name as PATHWAYPOINT, id as PATHWAYPOINTID FROM pathwayPoint", db.conn)
# 		landCover_df = pd.read_sql_query("SELECT name as LANDCOVER, id as LANDCOVERID FROM landCover", db.conn)


# 	months = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']
# 	detection = ['DETECTIONRATE','TREATMENTRATEFORUNDETECTED','TREATMENTEFFICACY']
# 	suitablity = ['FAVOURABLE','SUITABLE','MARGINAL','UNSUITABLE']
# 	conseq = ['ECONESTAB','ECONSPREAD','ENVIRONESTAB','ENVIRONSPREAD','HEALTHESTAB','HEALTHSPREAD','SOCIALESTAB','SOCIALSPREAD']

# 	df = pd.read_csv(driver)
# 	df = df[['CARRIER', 'SUBITEM','ITEM', 'SOURCE']]
# 	df = df.merge(pathwayPoint_df, how='left', on='ITEM')
# 	for var in months+detection+['SD']+suitablity:
# 		df[var] = None
# 	for var in ['SUITABILITY']+conseq:
# 		landCover_df[var] = None
# 	df.columns = [col.upper() for col in df.columns]
# 	landCover_df.columns = [col.upper() for col in landCover_df.columns]


# 	base_cols = {
# 		'carrier': ['CARRIER', 'SUBITEM','ITEM', 'ITEMID', 'SOURCE'],
# 		'units': ['SUBITEM','ITEM', 'ITEMID', 'SOURCE'] + months,
# 		'carrierRate': ['CARRIER', 'SUBITEM', 'ITEM', 'ITEMID', 'SOURCE'] + months,
# 		'carrierInfectionRate': ['CARRIER', 'SUBITEM', 'ITEM', 'ITEMID', 'SOURCE'] + months,
# 		'carriersPerUnit': ['CARRIER', 'SUBITEM', 'ITEM', 'ITEMID'] + months,
# 		'preborderDetection': ['CARRIER', 'SUBITEM', 'ITEM', 'ITEMID', 'SOURCE'] + detection,
# 		'pathwayDetection': ['CARRIER', 'ITEM', 'ITEMID', 'PATHWAYPOINT', 'PATHWAYPOINTID'] + detection,
# 		'carrierDailyMortalityRate': ['CARRIER', 'ITEM', 'ITEMID', 'PATHWAYPOINT', 'PATHWAYPOINTID'] + months,
# 		'pathogenDailyMortalityRate': ['CARRIER', 'ITEM', 'ITEMID', 'PATHWAYPOINT', 'PATHWAYPOINTID'] + months,
# 		'carrierDailyExitRate': ['CARRIER', 'ITEM', 'ITEMID', 'PATHWAYPOINT', 'PATHWAYPOINTID'] + months,
# 		'carrierDispersal': ['CARRIER', 'SD'],
# 		'transmissionRate': ['CARRIER'] + suitablity,
# 	}

# 	landCover_cols = {
# 		'landSuitability': ['LANDCOVER','LANDCOVERID','SUITABILITY'],
# 		'consequences': ['LANDCOVER','LANDCOVERID']+conseq
# 	}




# 	dataframes = {
# 		name: df[cols].drop_duplicates().reset_index(drop=True)
# 		for name, cols in base_cols.items()
# 	}

# 	dataframes.update({
# 		'hostMortalityRate': pd.DataFrame(columns=suitablity),
# 		'establishmentRate': pd.DataFrame(columns=suitablity),
# 		'establishmentDetection': pd.DataFrame(columns=detection),
# 		'establishmentMortalityRate': pd.DataFrame(columns=suitablity),
# 		'spreadRate': pd.DataFrame(columns=suitablity)
# 	})

# 	dataframes.update({
# 		name: landCover_df[cols].drop_duplicates().reset_index(drop=True)
# 		for name, cols in landCover_cols.items()
# 	})

# 	wb = Workbook()
# 	wb.remove(wb.active)

# 	for sheet_name, df in dataframes.items():
# 		ws = wb.create_sheet(title=sheet_name)
# 		ws.append(['ID', 'SCENARIOID']+df.columns.tolist())
# 		for row in df.itertuples(index=False):
# 			ws.append([None, None]+list(row))

# 	wb.save("output.xlsx")
