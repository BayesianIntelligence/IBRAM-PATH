import _env, cherrypy, webbrowser, json, logging, re, os, glob, psutil, signal, subprocess, tempfile, zipfile

import utils
from bidb import DB
from htm import node as n, toHtml, RawHtml

from openpyxl import load_workbook
from io import BytesIO

import pandas as pd
import numpy as np

serverDb = utils.serverDb
settings = utils.getSettings()


def runTemplate(strng, **kwargs):
	def templateReplace(m):
		field = m.group(1)
		if field in kwargs:
			if isinstance(kwargs[field],list):
				return "".join(str(v) for v in kwargs[field])
			else:
				return str(kwargs[field])
		return ''
	return re.sub(r'\[\[([^\]]*)\]\]', templateReplace, strng)

def runFileTemplate(fn, **kwargs):
	with open(fn, 'r', encoding='utf-8') as fl:
		return runTemplate(fl.read(), **kwargs)
	
class RequestUtils:
	def makePage(cls, templateFn = 'basic.html', **kwargs):
		return runFileTemplate(templateFn, **kwargs)


class IBRAMServer(RequestUtils):
	
	@cherrypy.expose	
	def getClimateMaps(self):
		with serverDb() as db:
			climateMaps = db.queryRows('SELECT * FROM climateMap')

		sortedMaps = sorted(climateMaps, key=lambda m: m['name'])
		maps = [['','','']]+[[m['fileName'], m['name']] for m in sortedMaps]
		
		return maps

	@cherrypy.expose
	def projectList(self):
		with serverDb() as db:
			projects = db.query('select id, name from project order by name')

			projectList = []
			for i,row in enumerate(projects):
				statusI = self.getProjectRunStatus(row['id'])
				projectList.append(n('div.project',
					n('a', href='/project?id={}'.format(row['id']), c=toHtml(row['name'])),
					n('span.runStatus', ['Needs Updates','Running Scenarios...','Updated'][statusI]),
				))
			
			return ''.join(str(n) for n in projectList)
		
	@cherrypy.expose
	def scenarioList(self, projectId):
		with serverDb() as db:
			scenarios = db.query('''select * from scenario
				where active = 1 and projectId = ?
				order by isBase desc, name''', [projectId])
			
			scenarioList = []
			for i,row in enumerate(scenarios):
				statusText, statusI = self.getScenarioRunStatus(row['id'])
				# print(row['id'], statusText, statusI)
				scenarioList.append(n('div.scenario'+('.isBase' if row['isBase'] else ''), dataId=row['id'], c=[
					n('span.name', n('a', href=f"/scenario?id={row['id']}", c=toHtml(row['name']))),					
					n('span.actions', c=[
						n('span.table', title="Table", onclick='scenarioTable({});'.format(row['id']), c=n('i', Class='fas fa-fw fa-table')),		
						n('span.run'+('.hide' if statusI == 1 else ''), title="Update", onclick='runScenario({});'.format(row['id']), c=n('i', Class="fas fa-fw fa-bolt")),
						n('span.cancel'+('.hide' if statusI != 1 else ''), title="Cancel Update", onclick='cancelRun({});'.format(row['id']), c=n('i', Class="fas fa-fw fa-times")),
						n('span.viewOutput'+('.hide' if statusI != 2 else ''), title="View Outputs", onclick='viewOutput({});'.format(row['id']), c=n('i', Class="fas fa-fw fa-chart-bar")),
						n('span.exportOutput'+('.hide' if statusI != 2 else ''), title="Export Outputs", onclick='exportOutput({});'.format(row['id']), c=n('i', Class="fas fa-fw fa-download")),
						n('span.delete', title="Delete", onclick='deleteScenario({});'.format(row['id']), c=n('i', Class='fas fa-fw fa-trash')),
						n('span.makeBase', title="Make Base Scenario", onclick=f'makeBase({row["id"]});', c=n('i', Class='fas fa-fw fa-star')) if not row['isBase'] else '',
					]),
					n('span.status', c=statusText)
				]))
			
			return ''.join(str(n) for n in scenarioList)
	
	def saveTable(self, table, tableValues, pk = 'id'):
		newPks = []
		with serverDb() as db:
			for row in tableValues['data']:
				rowUpdates = dict(zip(tableValues['fields'], row))
				newPk = db.replace(table, rowUpdates, pk)#, fields)
				newPks.append(newPk)
		
		return newPks


	@cherrypy.expose
	def saveScenario(self, scenarioTables):
		with serverDb() as db:
			scenarioTables = json.loads(scenarioTables)
			for name, table in scenarioTables.items():
				self.saveTable(name, table)			
		
		return json.dumps({'ok': True})

	
	@cherrypy.expose
	def saveParameters(self, paramTables):
		paramTables = json.loads(paramTables)

		with serverDb() as db:
			class _Save:

				def climateMaps(_self, tableValues):
					fields = tableValues['fields']
					data_rows = [dict(zip(fields, row)) for row in tableValues['data']]

					# Clear the entire table first
					db.query("DELETE FROM climateMap")

					# Insert all provided rows (without id)
					for row in data_rows:
						db.query("INSERT INTO climateMap (name, fileName) VALUES (?, ?)", (row['name'], row['fileName']))
						

				def items(_self, tableValues):
					
					fields = tableValues['fields']
					data_rows = [dict(zip(fields, row)) for row in tableValues['data']]

					# Clear the entire table first
					db.query("DELETE FROM item")

					# Insert all provided rows (without id)
					for row in data_rows:
						db.query("INSERT INTO item (name) VALUES (?)", [row['name']])
					

				def pathwayPoints(_self, tableValues):
					fields = tableValues['fields']
					data_rows = [dict(zip(fields, row)) for row in tableValues['data']]
					
					item_rows = db.query("SELECT id, name FROM item")
					item_lookup = {row['name']: row['id'] for row in item_rows}

					# Clear the entire table first
					db.query("DELETE FROM pathwayPoint")

					# Insert all provided rows (without id)
					for row in data_rows:
						db.query(
							"INSERT INTO pathwayPoint (name, item, itemId, tableName, shape, timeAtSite) VALUES (?, ?, ?, ?, ?, ?)",
							[row['name'], row['item'], item_lookup.get(row['item']), row['tableName'], row['shape'], row['timeAtSite']]
						)

			save = _Save()

			for name, table in paramTables.items():
				print(name)
				getattr(save, name)(table)

			return json.dumps({'ok': True})

		
	@cherrypy.expose
	def downloadParameterData(self, file):
		if file.endswith('.csv'):
			for subdir in ['pathway', 'climateMaps']:
				full_path = os.path.join(settings['inputsDir'], subdir, file)
				if os.path.exists(full_path):
					cherrypy.response.headers['Content-Type'] = 'text/csv'
					cherrypy.response.headers['Content-Disposition'] = f'attachment; filename="{file}"'
					with open(full_path, 'rb') as f:
						return f.read()
		raise cherrypy.HTTPError(404, 'File not found')


	@cherrypy.expose
	@cherrypy.tools.allow(methods=['POST'])
	def uploadNewParameterData(self, file, fileName):
		upload = file.file.read()
		for subdir in ['pathway', 'climateMaps']:
			file_path = os.path.join(settings['inputsDir'], subdir, fileName)
			if os.path.exists(file_path):
				with open(file_path, 'wb') as f:
					f.write(upload)
				return 'OK'
		raise cherrypy.HTTPError(404, 'Destination file not found')

	@cherrypy.expose
	def viewParameterData(self, table, id):
			with serverDb() as db:
				data = db.queryRow(f'SELECT * FROM {table} WHERE id = ?', [id])
				
			if table == 'pathwayPoint':
				tableName = data['tableName']
				inputTableFile = os.path.join(settings['inputsDir'], 'pathway', tableName+'.csv')
			if table == 'climateMap':
				tableName = data['fileName']
				inputTableFile = os.path.join(settings['inputsDir'], 'climateMaps', tableName+'.csv')		
				
			inputFileName = os.path.basename(inputTableFile) 

			head = [
			n('script', '''
				window.addEventListener('DOMContentLoaded', event => {
					$('.trail a').text('Close').attr('href', 'javascript:window.close()');
				});
			''')]
			body = [
				n('p.controls',
	  				n('button', type='button', c='Download Data', onclick=f"downloadParameterData('{inputFileName}')"),
					n('button', type='button', c='Upload New Data', onclick=f"uploadNewParameterData('{inputFileName}')"),
				),
			]
				
			df = pd.read_csv(inputTableFile)
			body.append(RawHtml(df.to_html(index=False, border=0))			)
			return runFileTemplate('basic.html', head = head, body = n('div.pageBody', body))


	@cherrypy.expose
	def uploadMap(self, type=None, file=None):
		print('here')
		if not type or not file:
			cherrypy.response.status = 400
			return {'error': 'Missing type or file'}

		upload_path = f'./inputs/{type}/{file.filename}'
		with open(upload_path, 'wb') as out:
			out.write(file.file.read())

		print(f"Uploaded: {file.filename} as {upload_path}")
		return json.dumps({'status': 'ok'})


	@cherrypy.expose
	def pathwayParams(self):
		with serverDb() as db:
			climateMaps = db.queryRows("""select * from climateMap""")
			climateMapHeaders = ['id', 'Climate Map Name', 'Data Source']			
			climateMapTabExtra = [
				n('div',
					n('button', type='button', c='Add Climate Map', onclick='addClimateMap()'),
					n('button', type='button', c='Delete Climate Map', onclick='deleteClimateMap()'),
				),
			]			

			items = db.queryRows("""select * from item""")
			itemHeaders = ['id', 'Item Name']			
			itemsTabExtra = [
				n('div',
					n('button', type='button', c='Add Item', onclick='addItem()'),
					n('button', type='button', c='Delete Item', onclick='deleteItem()'),
				),
			]

			pathwayPoints = db.queryRows("""select * from pathwayPoint""")
			pathwayPointsHeaders = ['id', 'Point Name', 'Item Name', 'itemId', 'Data Source', 'Type', 'Time At Site']
			pathwayPointsTabExtra = [
				n('div',
					n('button', type='button', c='Add Pathway Point', onclick='addPathwayPoint()'),
					n('button', type='button', c='Delete Pathway Point', onclick='deletePathwayPoint()'),
				),
			]

			sheetsData = [
				{'id': 'climateMaps', 'tabName': 'Climate Maps', 'table': climateMaps, 'headerLabels': climateMapHeaders, 'tabExtra': climateMapTabExtra},
				{'id': 'items', 'tabName': 'Items', 'table': items, 'headerLabels': itemHeaders, 'tabExtra': itemsTabExtra},
				{'id': 'pathwayPoints', 'tabName': 'Pathway Points', 'table': pathwayPoints, 'headerLabels': pathwayPointsHeaders, 'tabExtra': pathwayPointsTabExtra},
			]

			tabButtons = n('div.tabButtons')
			tabSheets = n('div.tabSheets')
			tabSet = n('div.tabSet',
				tabButtons,
				tabSheets,
			)
			
			def handleCellVal(val):
				return re.sub(r'^(.*?)\{\{|\}\}(.*?)\{\{|\}\}(.*?)$', lambda m: toHtml("".join([f for f in m.groups() if f is not None])), str(val))
			
			for sheetData in sheetsData:
				tabButtons.append(
					n('button', type='button', dataFor=sheetData['id'], c=sheetData['tabName'])
				)
				tabSheets.append(
					n('div.tabSheet',
						n('table.dataTable',
							dataName = sheetData['id'],
							data = sheetData['table'],
							headerLabels = sheetData['headerLabels'],
							hidden=['id', 'itemId'],
							readonly = ['actions'],
							omit = [],
							handlers = {
								'tableName': 'textarea',
							},
							defaultDataCellHandler = lambda val, rowI, cellI: n('td', n('div', RawHtml(handleCellVal(val))))
						) if sheetData.get('table') else None,
						sheetData.get('tabExtra', ''),
						dataName=sheetData['id']
					)
				)
			
			inputFolders = [os.path.basename(os.path.dirname(f)) for f in glob.glob(settings['inputsDir'] + '*/')]
			
			return n('div.params',
				n('style', """
				.params td.include { text-align: center; }
				.params td { height: 1px; }
				@-moz-document url-prefix() { .params td { height: 100%; } }
 				.params td div[contenteditable] { position: relative; height: 100%; width: 100%;
				   outline: dotted 1px gray; display: inline-table; }
				tr.new { background: rgb(255,255,204);  }
				.dialog .error { font-size: 0.8em; color: red; }
				.dialog [name=fileName] { width: 30em; }
				"""),
				n('script', """
				var inputFolders = """+json.dumps(inputFolders)+""";
				
				
				$().ready(function() {
					addViewLinks()	
				
					o = {}
					o.activeSheet = o.activeSheet || 'climateMaps';
					document.querySelector(`.params .tabButtons button[data-for=${o.activeSheet}]`).click();
					
				});
				"""),
					n('h2', 'Parameters'),
					RawHtml(tabSet),
					n('div.controls', n('button', type='button', onclick='saveParameters()', c='Save')),			
				).str()	


		
	outputPrefixMap = {
		'cs': 'CS_',
		'ls': 'LS_',
		'habitat': 'Habitat_Suitability_',
		'exposure': 'Exposure_Pests_Density_',
		'disperse': 'Disperse_Pests_Density_',
		'accumulated': 'x_Pests_',
		'establishment': 'x_Establishment__',
		'spread': 'x_Spread__',
		'consequencesEconomic': 'Economic_Consequences_',
		'consequencesEnvironmental': 'Environmental_Consequences_',
		'consequencesHealth': 'Human_Health_Consequences_',
		'consequencesSocial': 'Social_Cultural_Consequences_',
	}

	@cherrypy.expose
	def getMapData(self, scenarioId, stage, month):

		def getScenarioMapData(scenarioId, stage, month):
			scenarioDir = utils.getOutputDir(scenarioId)
			month = int(month)
			
			outputPrefix = self.outputPrefixMap[stage]
			fn = os.path.join('inputs', '10kmHexClippedNZTM', '10kmHexClippedNZTM.csv')
			df_hex = pd.read_csv(fn)
			
			fn = os.path.join(scenarioDir, '{stage}{month}.csv'.format(stage=outputPrefix,month=month))
			df = pd.read_csv(fn)
			
			df = pd.merge(df, df_hex, on='Code', how='inner')
			
			df = df[df['area_sqkm'] > 20]
		
			mapData = df.set_index('Code')[['uPeSqKm', 'EA_Name']].to_dict(orient='index')
			return mapData
				

		with serverDb() as db:
			scenario = db.queryRow('SELECT projectId FROM scenario WHERE id = ?', [scenarioId])
			baseScenarioId = db.queryValue('select id from scenario where projectId = ? and isBase', [scenario['projectId']])
			

		mapData = getScenarioMapData(scenarioId, stage, month)
		baseMapData = getScenarioMapData(baseScenarioId, stage, month)
		
		cherrypy.response.headers['Content-Type'] = 'application/json'
		return json.dumps({
			'mapData': mapData,
			'baseMapData': baseMapData
		}).encode('utf-8')
	
	@cherrypy.expose
	def getTimeData(self, scenarioId, stage, locationCode):
		with serverDb() as db:
			scenario = db.queryRow('SELECT projectId FROM scenario WHERE id = ?', [scenarioId])
			project = db.queryRow('SELECT runLength FROM project WHERE id = ?', [scenario['projectId']])
			
		scenarioDir = utils.getOutputDir(scenarioId)
		outputPrefix = self.outputPrefixMap[stage]
		numMonths = project['runLength']	
		locationCode = int(locationCode)
			
		timeData = []
		
		for month in range(numMonths):
			fn = os.path.join(scenarioDir, '{stage}{month}.csv'.format(stage=outputPrefix,month=month))
			df = pd.read_csv(fn)
			
			value = df.loc[df['Code'] == locationCode, 'uPeSqKm'].values[0]
			
			timeData.append(value)
			
		# print(timeData)
		
		return json.dumps(timeData)


	def getPlaces(self):
		fn = os.path.join('inputs', '10kmHexClippedNZTM', '10kmHexClippedNZTM.csv')
		df = pd.read_csv(fn)	
		places = df.set_index('Code')['EA_Name'].to_dict()
		return places
	



	@cherrypy.expose
	def downloadScenario(self, scenarioId):
		with serverDb() as db:
			scenario = db.queryRow('SELECT projectId FROM scenario WHERE id = ?', [scenarioId])
			project = db.queryRow('SELECT name FROM project WHERE id = ?', [scenario['projectId']])
			fileName = f"{project['name']}.xlsx"
	
			tables = [
				'vector', 'units', 'infestationRate', 'vectorInfectionRate', 'itemInfectionRate',
				'itemVectorTransmissionRate', 'preBorderVectorDetection', 'preBorderItemDetection',
				'vectorsPerUnit', 'vectorPathwayDetection', 'itemPathwayDetection', 'vectorDailyEscapeRate',
				'itemDailyEscapeRate', 'vectorDailyMortalityRate', 'itemDailyMortalityRate',
				'vectorTransmissionRate', 'mortalityRate', 'establishmentRate', 'spreadRate',
				'eradicationRate', 'landSuitability', 'consequences', 'eradicationDetection'
			]
					
			with BytesIO() as output:
				with pd.ExcelWriter(output, engine='openpyxl') as writer:
					for table in tables:
						query = f"SELECT * FROM {table} WHERE scenarioId = ?"
						df = pd.read_sql_query(query, db.conn, params=[scenarioId])
						df.to_excel(writer, sheet_name=table, index=False)
				
				excel_data = output.getvalue()

			cherrypy.response.headers['Content-type'] = 'application/vnd.ms-excel'
			cherrypy.response.headers['Content-disposition'] = f'attachment; filename={fileName}'

			return excel_data

	
	@cherrypy.expose
	def deleteScenario(self, scenarioId):
		print('deleting scenario', scenarioId)
		returnVal = 'OK'
		with serverDb() as db:
			isBase = db.queryValue('select isBase from scenario where id = ?', [scenarioId])
			projectId = db.queryValue('select projectId from scenario where id = ?', [scenarioId])
			db.query('delete from scenario where id = ?', [scenarioId])
			
			tables = [
				'vector', 'units', 'infestationRate', 'vectorInfectionRate', 'itemInfectionRate',
				'itemVectorTransmissionRate', 'preBorderVectorDetection', 'preBorderItemDetection',
				'vectorsPerUnit', 'vectorPathwayDetection', 'itemPathwayDetection', 'vectorDailyEscapeRate',
				'itemDailyEscapeRate', 'vectorDailyMortalityRate', 'itemDailyMortalityRate',
				'vectorTransmissionRate', 'mortalityRate', 'establishmentRate', 'spreadRate',
				'eradicationRate', 'landSuitability', 'consequences', 'eradicationDetection'
			]

			for table in tables:
				db.query(f"delete from {table} where scenarioId = ?", (scenarioId,))

			
			if isBase == 1:
				returnVal = 'DeleteProject'
				db.query('delete from project where id = ?', [projectId])
				scenarios = db.query('select * from scenario where active = 1 and projectId = ?', [projectId])
				for scenario in scenarios:
					self.deleteScenario(scenario['id'])
		
		return returnVal
	

	@cherrypy.expose
	def copyBaseScenario(self, baseScenarioId):
		print(f'Copying scenario {baseScenarioId}')
		with serverDb() as db:
			source = db.queryRow('select * from scenario where id = ?', [baseScenarioId])
			
			db.query('insert into scenario (name, projectId, isBase, active) values (?,?,?,?)', ['Base Copy', source['projectId'], False, True])
			scenarioId = db.queryValue('select last_insert_rowid()')

			tables = [
				'vector', 'units', 'infestationRate', 'vectorInfectionRate', 'itemInfectionRate',
				'itemVectorTransmissionRate', 'preBorderVectorDetection', 'preBorderItemDetection',
				'vectorsPerUnit', 'vectorPathwayDetection', 'itemPathwayDetection', 'vectorDailyEscapeRate',
				'itemDailyEscapeRate', 'vectorDailyMortalityRate', 'itemDailyMortalityRate',
				'vectorTransmissionRate', 'mortalityRate', 'establishmentRate', 'spreadRate',
				'eradicationRate', 'landSuitability', 'consequences', 'eradicationDetection'
			]

			for table in tables:
				rows = db.query(f"select * from {table} where scenarioId = ?", [baseScenarioId])
				
				for row in rows:
					row = dict(row)
					row['id'] = -1
					row['scenarioId'] = scenarioId
					
					db.replace(table, row, pk='id')

			return str(scenarioId)
				
			
	@cherrypy.expose
	def addScenario(self, wb, scenarioId):
		print('Adding scenario', scenarioId)
		for s in wb:
			scenarioid_col = None
			id_col = None

			for col in range(1, s.max_column + 1):
				header = s.cell(row=1, column=col).value
				if header:
					if header.lower() == 'scenarioid':
						scenarioid_col = col
					elif header.lower() == 'id':
						id_col = col

			if scenarioid_col:
				for row in range(2, s.max_row + 1):
					has_data = any(
						s.cell(row=row, column=c).value not in (None, '')
						for c in range(1, s.max_column + 1)
						if c != scenarioid_col
					)
					if has_data:
						s.cell(row=row, column=scenarioid_col).value = scenarioId
						if id_col:
							s.cell(row=row, column=id_col).value = None

		with serverDb() as db:
			db.loadDataFromExcel(wb)
			
	@cherrypy.expose
	def uploadScenario(self, projectId, sheet):
		print('uploading scenario', projectId, sheet.filename)
		# return 1
		base,ext = os.path.splitext(sheet.filename)
		if ext != '.xlsx':
			errorDetails = {}
			errorDetails["type"] = 'error'
			errorDetails["message"] = 'Bad file type, must be .xlsx'
			return json.dumps(errorDetails)
			
		with serverDb() as db:
			name, _ = os.path.splitext(sheet.filename)
			db.query('insert into scenario (name, projectId, isBase, active) values (?,?,?,?)', [name, projectId, False, True])
			scenarioId = db.queryValue('select last_insert_rowid()')
		
		wb = load_workbook(sheet.file)
		self.addScenario(wb, scenarioId)
		
		return str(scenarioId)


	@cherrypy.expose
	def addProject(self, sheet):
		print('Adding project', sheet.filename)
		base,ext = os.path.splitext(sheet.filename)
		if ext != '.xlsx':
			errorDetails = {}
			errorDetails["type"] = 'error'
			errorDetails["message"] = 'Bad file type, must be .xlsx'
			return json.dumps(errorDetails)
			
		with serverDb() as db:
			projectId = db.replace('project', {'id': -1, 'name': base, 'burnin': 12, 'runlength': 24, 'dispersalsd': 3.33}, 'id')
			db.query('insert into scenario (name, projectId, isBase, active) values (?,?,?,?)', ['Base', projectId, True, True])
			scenarioId = db.queryValue('select last_insert_rowid()')
		
		wb = load_workbook(sheet.file)
		self.addScenario(wb, scenarioId)
		
		return str(projectId)

		


	@cherrypy.expose
	def projectNameUpd(self, id, name):
		print('updating project name', id, name)
		with serverDb() as db:
			db.query('update project set name = ? where id = ?', [name, id])
			
			return "OK"
	
	@cherrypy.expose
	def scenarioNameUpd(self, id, name):
		print('updating scenario name', id, name)
		with serverDb() as db:
			db.query('update scenario set name = ? where id = ?', [name, id])
			
			return "OK"
		

	@cherrypy.expose
	def getProjectRunStatus(self, projectId):
		with serverDb() as db:
			scenarios = db.query('SELECT * from scenario where projectId = ?', [projectId])
			status = 2
			for scenario in scenarios:
				if scenario['complete'] is not None:
					status = min(status, 2)
				elif scenario['processId'] is not None and psutil.pid_exists(scenario['processId']):
					status = min(status, 1)
				else:
					status = min(status, 0)
		return status
		
	@cherrypy.expose
	def getScenarioRunStatus(self, scenarioId):
		with serverDb() as db:
			scenario = db.queryRow('SELECT * from scenario where id = ?', [scenarioId])
		
		if scenario['complete'] is not None:
			return 'Updated', 2
		elif scenario['processId'] and psutil.pid_exists(int(scenario['processId'])):
			return scenario['status'], 1
		
		return 'Needs Update', 0

	@cherrypy.expose
	def runScenario(self, scenarioId):
		print('running scenario',scenarioId)
		with serverDb() as db:
			self.killScenario(scenarioId)
			process = subprocess.Popen(['python', '-u', 'run_model.py', '--id', str(scenarioId)])
			db.replace('scenario', {'complete': None, 'id': scenarioId, 'processId': process.pid, 'status': 'Running'}, 'id')


	@cherrypy.expose
	def runProject(self, projectId):
		print('running project',projectId)
		with serverDb() as db:
			scenarios = db.query('''select scenario.id from scenario where projectId = ? and complete is null''', [projectId])
			for scenario in scenarios:
				self.runScenario(scenario['id'])
	


	@cherrypy.expose
	def killScenario(self, scenarioId):
		with serverDb() as db:
			pid = db.queryValue('select processId from scenario where id = ?', [scenarioId])
			if pid is not None and psutil.pid_exists(pid):
				print('killing scenario', scenarioId, pid)
				try:
					proc = psutil.Process(pid)
					proc.terminate()
					proc.wait(timeout=5)
				except Exception as e:
					print('Failed to kill or wait for process:', e)

	@cherrypy.expose		
	def resetScenario(self, scenarioId):
		print('resetting scenario',scenarioId)
		with serverDb() as db:
			self.killScenario(scenarioId)
			db.replace('scenario', {'complete': None, 'id': scenarioId}, 'id')



			
	@cherrypy.expose		
	def resetProject(self, projectId):
		print('resetting project',projectId)
		with serverDb() as db:
			scenarios = db.query('''select scenario.id from scenario where projectId = ?''', [projectId])
			for scenario in scenarios:
				self.resetScenario(scenario['id'])
		
			
	@cherrypy.expose
	def generateInput(self):
		pass

	@cherrypy.expose
	def index(self):
		projectList = self.projectList()
		
		head = [
			n('script', """
				$(document).ready(function() {
					setInterval(function() {
						$.get("/projectList", function(data) {
							$(".projectsList").html(data);
						});
					}, 5000);
	
					$('button.manageParams').on('click', function() {
						window.location.href='/parameters'
					});
	 
					$('button.addProject').on('click', function() {
						ui.dialog(`<h2>Add Project</h2>
							<p>Would you like to create a new project or upload an existing spreadsheet?`, {
							buttons: [
								$('<button>').text('Generate').on('click', function() {
									//window.location.href='/generateInput'
									ui.dismissDialogs();
								}),
								$('<button>').text('Upload').on('click', function() {
									$('input.addProjectFile').val('');
									$('input.addProjectFile').click();
									ui.dismissDialogs();
								}),
								$('<button>').text('Cancel').on('click', ui.dismissDialogs),
							],
						});
					});
	
					$('input.addProjectFile').on('change', function() {
							var f = new FormData();
							f.append('sheet', this.files[0], this.files[0].name);
							$.post({
								url: '/addProject',
								data: f,
								success: (data) => {
									data = JSON.parse(data);
									if(data.type == 'error') {
										dismissDialogs();
										ui.message(data.message);
									}
									else {
										window.location.href = '/project?editName=1&id='+data;
									}
								},
								contentType: false,
								processData: false,
							});
						});
					});
			"""),
		]
		
		body = [
			n('div.mainPage',
				n('div.projects',
					n('h2', n('i', Class='fas fa-fw fa-bug'), RawHtml('&nbsp;Projects')),
					n('div.projectsList', RawHtml(projectList)),
					n('div.controls', [
						n('button.addProject', type='button', c='Add Project'),
						n('button.manageParams', type='button', c='Manage Pathway Parameters'),
						n('input.addProjectFile', type='file', style='display:none'),
					]),
				),
				),
			]
		return runFileTemplate('basic.html', head = head, body = body)
	


	@cherrypy.expose
	def parameters(self):
		with serverDb() as db:
			
			head = [
				n('script', '''
					$(document).ready(function() {	
					});
				''')
			]
		
			body = [
				n('div.mainPage',
					n('div.parameters.contentPanel',
						RawHtml(self.pathwayParams())
					)
				)
			]
			
			return runFileTemplate('basic.html', head = head, body = body, trail = [
				n('li', RawHtml('<i class="fas fa-fw fa-sticky-note"></i> parameters')),
			])


	@cherrypy.expose
	def saveSettings(self, **kwargs):
		print(kwargs)
		with serverDb() as db:
			db.replace('project', kwargs, 'id')
		return "OK"

	@cherrypy.expose
	def project(self, id, editName = False):
		with serverDb() as db:
			project = db.queryRow('select * from project where id = ?', [id])	
			baseScenarioId = db.queryValue('select id from scenario where projectId = ? and isBase', [id])
			

			scenarioList = self.scenarioList(id)
			maps = self.getClimateMaps()
			
			# print(maps)
			
			head = [
				n('script', '''
	  
	  				var projectId = '''+id+''';
	  				var baseScenarioId = '''+str(baseScenarioId)+''';		
	  
					$(document).ready(function() {
						setInterval(function() {
							updateScenarioList()
						}, 10000);
					});
				''')
			]
			
			
			body = [
				n('div.scenariosPage',
					n('div.scenariosPanel.panel',
						n('h2', dataFieldEdit=f'/projectNameUpd?id={id}&name=', c=n('span.value', toHtml(project['name']))),
						n('div.scenarios', n('div.scenarioList', RawHtml(scenarioList))),
						n('div.controls', [
							n('button.addScenario', type='button', onclick='addScenario()', c='Add Scenario'),
							n('input.addScenarioFile', type='file', style='display:none'),
							n('button.projectUpdate', type='button', onclick='runProject()', c='Update All'),
						]),
					),
					n('div.config.panel', 
						n('h2', 'Settings'),
						n('div.settingsPanel',
							n('form.grid.settings', 
								n('label', 'Burn-in months', title="Number of months for burn-in process"), 
								n('input', type='text', name="burnin", value=project['burnIn']),
								n('label', 'Run months', title="Number of months for the run itself"), 
								n('input', type='text', name="runlength", value=project['runLength']),
								n('label', 'Climate Map', title="Climate map to be used"), 
								n('select', data=maps, name="climateMap",  selected=project['climateMap']),
								n('label', 'Dispersal (Std Dev)', title="Standard deviation distance in meters for dispersal after exposure"), 
								n('input', type='text', name="dispersalSd", value=project['dispersalSd']),
								n('input', type='hidden', name="id", value=id),
							),
						),
						n('div.controls', [
						       n('button.saveSettings', type='button', onclick="saveSettings($('.settings'))", c='Save Settings'),
						]),
						       
					),
				),
			]
			
		return runFileTemplate('basic.html', head = head, body = body, trail = [n('li',  n('i', Class='fas fa-fw fa-bug'), RawHtml('&nbsp;Projects'))])


		


	@cherrypy.expose
	def scenarioTable(self, id):

		with serverDb() as db:
			scenario = db.queryRow('select projectId from scenario where id = ?', [id])
			baseScenarioId = db.queryValue('select id from scenario where projectId = ? and isBase', [scenario['projectId']])	

			tables = [
				'vector', 'units', 'infestationRate', 'vectorInfectionRate', 'itemInfectionRate',
				'itemVectorTransmissionRate', 'preBorderVectorDetection', 'preBorderItemDetection',
				'vectorsPerUnit', 'vectorPathwayDetection', 'itemPathwayDetection', 'vectorDailyEscapeRate',
				'itemDailyEscapeRate', 'vectorDailyMortalityRate', 'itemDailyMortalityRate',
				'vectorTransmissionRate', 'mortalityRate', 'establishmentRate', 'spreadRate',
				'eradicationRate', 'landSuitability', 'consequences', 'eradicationDetection'
			]
			
			sheetsData = []

			for table in tables:
				table_data = db.queryRows(f'SELECT * FROM {table} WHERE scenarioId = ?', [id])
				base_table_data = db.queryRows(f'SELECT * FROM {table} WHERE scenarioId = ?', [baseScenarioId])
				columns = db.query(f"PRAGMA table_info('{table}')").fetchall()
				header_labels = [column[1].upper() for column in columns]
				
				sheetsData.append({
					'id': table,
					'tabName': table,
					'table': table_data,
					'baseTable': base_table_data,
					'headerLabels': header_labels,
					'tabExtra': []
				})
				# if table == 'preBorderVectorDetection':
				# 	print(sheetsData)

			tabButtons = n('div.tabButtons')
			tabSheets = n('div.tabSheets')
			tabSet = n('div.tabSet',
				tabButtons,
				tabSheets,
			)
						
			def handleCellVal(val, baseVal):
				diff = val != baseVal
				attrs = {}
				if diff:
					attrs['style'] = 'background-color:#fdd;'
					attrs['title'] = f'Base: {baseVal}'
				return n('div', toHtml(val), **attrs)

			for sheetData in sheetsData:
				tabButtons.append(
					n('button', type='button', dataFor=sheetData['id'], c=sheetData['tabName'])
				)
				tabSheets.append(
					n('div.tabSheet',
						# Only show if table is set
						n('table.dataTable',
							dataName = sheetData['id'],
							data = sheetData['table'],
							headerLabels = sheetData['headerLabels'],
							hidden=['id', 'scenarioId', 'itemId', 'vectorId', 'pathwayPointId'],
							# readonly = ['item','vector'],
							# omit = [],
							defaultDataCellHandler = lambda val, rowI, cellI: n('td', RawHtml(handleCellVal(val, sheetData['baseTable'][rowI][list(sheetData['table'][0].keys())[cellI]])))
						) if sheetData.get('table') else None,
						sheetData.get('tabExtra', ''),
						dataName=sheetData['id']
					)
				)
				
			return n('div.params',
				n('style', """
				.params td.include { text-align: center; }
				.params td { height: 1px; }
				@-moz-document url-prefix() { .params td { height: 100%; } }
 				.params td div[contenteditable] { position: relative; height: 100%; width: 100%;
				   outline: dotted 1px gray; display: inline-table; }
				tr.new { background: rgb(255,255,204);  }
				.dialog .error { font-size: 0.8em; color: red; }
				.dialog [name=fileName] { width: 30em; }
				"""),
				n('script', """
				
				function setupParameterEditing(o = {}) {\
					o.activeSheet = o.activeSheet || 'units';
					$('.tabSheet td div').toArray().map(el => {
						let mgr = tableManager(el.closest('table'));
						let readOnlys = ['item','source','subItem', 'vector','name','pathwayPoint','landcover'];
						if ( !readOnlys.includes(mgr.fieldIds[el.closest('td').cellIndex]) ) {
							el.setAttribute('contenteditable', 'true');
						}
					});
				}
				
				$().ready(function() {
					setupParameterEditing();
				});
				"""),
				n('div', RawHtml(tabSet)),
				).str()	

	@cherrypy.expose
	def saveBaseScenario(self, scenarioId):
		with serverDb() as db:
			scenario = db.queryRow('select projectId from scenario where id = ?', [scenarioId])
			projectId = scenario['projectId']
			db.query('update scenario set isBase = 0 where projectId = ? and id <> ?', [projectId, scenarioId])
			db.query('update scenario set isBase = 1 where projectId = ? and id = ?', [projectId, scenarioId])
			return "OK"

	@cherrypy.expose
	def scenario(self, id, editName = False, table = 0):
		print('loading scenario page', id)
		with serverDb() as db:
			scenario = db.queryRow('''select * from scenario where id = ?''', [id])

			head = [
				n('script', '''
	  			var scenarioId = '''+json.dumps(id)+''';
				var tableIndex = '''+json.dumps(table)+''';
				
				$(document).ready(function() {				
					$(`.tabButtons button:nth(${tableIndex})`).click();
					
					$('button.download').on('click', function() {		
						window.location.href = '/downloadScenario?scenarioId='+scenarioId;
					});
					
					$('button.saveScenario').on('click', function() {
						console.log('here')
						saveScenario(function(data) {
							if(data=='OK') {
								window.location.href = '/scenario?id='+scenarioId+'&table='+($('.tabButtons .selected').index()-1);
							}
						});
					});
					
					$('.tabSet')[0].addEventListener('input', (event) => {
						$('button.saveScenario').attr('disabled', null);
					}, true);
				});
				'''),
			]
			
			body = [
				n('h2', dataFieldEdit='/scenarioNameUpd?id={}&name='.format(id), c=n('span.value',toHtml(scenario['name']))),
				n('div.controls', [
					n('button.saveScenario', disabled='disabled', type='button', c='Save'),
					n('button.download', type='button', c='Download'),
				]),
				self.scenarioTable(id)
			]
			
			return runFileTemplate('basic.html', head = head, body = body, trail = [
				n('li',n('a', href='/project?id={}'.format(scenario['projectId']), c=[n('i', Class="fas fa-fw fa-bug"), 'Project'])),
				n('li', n('i', Class="fas fa-fw fa-edit"), 'Scenario'),
			])
		
	@cherrypy.expose
	def exportOutput(self, scenarioId, token = None):
		t = tempfile.TemporaryFile()
		outputDir = utils.getOutputDir(scenarioId)
		with zipfile.ZipFile(t, 'w') as zip:
			for fn in os.listdir(outputDir):
				if fn.endswith('.csv'):
					print('zipping', fn)
					zip.write(os.path.join(outputDir, fn), os.path.split(fn)[1])

		cherrypy.response.cookie['fileDownloadToken'] = token
		cherrypy.response.headers['Content-Type'] = 'application/zip'
		cherrypy.response.headers['Content-Disposition'] = f'attachment; filename="scenario{scenarioId}_export.zip"'
		
		t.seek(0)
		return cherrypy.lib.file_generator(t)
		
	

	@cherrypy.expose
	def getTableData(self, scenarioId, tableDiff):
		val_cols = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

		outputDir = f'outputs/scenario{scenarioId}'
		outCsvFn = os.path.join(outputDir, "Entries.csv")
		df = pd.read_csv(outCsvFn)
		df.columns = [col.upper() for col in df.columns]
		
		if tableDiff == '1':
			with serverDb() as db:
				scenario = db.queryRow('select projectId from scenario where id = ?', [scenarioId])
				baseScenarioId = db.queryValue('select id from scenario where projectId = ? and isBase', [scenario['projectId']])
			
			base_outputDir = f'outputs/scenario{baseScenarioId}'
			base_outCsvFn = os.path.join(base_outputDir, "Entries.csv")
			base_df = pd.read_csv(base_outCsvFn)
			base_df.columns = [col.upper() for col in base_df.columns]
				
			diff_df = df.copy()
			diff_df[val_cols] = df[val_cols] - base_df[val_cols]
			df = diff_df


		df = df.drop(columns=['SCENARIOID','ITEMID','VECTORID'])
		df = df[['ITEM', 'SUBITEM', 'SOURCE','VECTOR']+val_cols]
		df = df.fillna('')
		
		vmin_raw = df[val_cols].min().min()
		vmax_raw = df[val_cols].max().max()
		abs_max = max(abs(vmin_raw), abs(vmax_raw))
		vmin, vmax = -abs_max, abs_max	
		cmap = 'bwr'

		styled_df = df.style.background_gradient(cmap=cmap, vmin=vmin, vmax=vmax, subset=val_cols)
		styled_df = styled_df.format(
			lambda x: '0' if x == 0 else f"{x:.5f}" if isinstance(x, float) else x
		)
		table = styled_df.hide(axis="index").render()
		

		return str(n('div.tableContainer', RawHtml(table)))
	

	@cherrypy.expose
	def outputTable(self, scenarioId, tableDiff=False):
		with serverDb() as db:
			scenario = db.queryRow('select isBase, projectId from scenario where id = ?', [scenarioId])
			isBase = scenario['isBase']			
		
		# print(isBase)
		
		tableHtml = self.getTableData(scenarioId, tableDiff)
		
		body = n('script', """	
			var scenarioId = """+json.dumps(scenarioId)+""";
			
			function toggleTableDifference(button) {
				button.dataset.diff = button.dataset.diff == '1' ? '0' : '1';
				
				if (button.dataset.diff == '1') {
					$(button).text('Hide Difference from Base');
				}
				else {
					$(button).text('Show Difference from Base');
				}

				const panel = document.querySelector('.resultsPanel');

				fetch(`/getTableData?scenarioId=${scenarioId}&tableDiff=${button.dataset.diff}`)
					.then(response => response.text())
					.then(html => {
						 document.querySelector('.tableContainer').outerHTML = html;
					});
			}
									
			$(document).ready(function() {
			
			});""")
		
		mainBody = n('div', [
			body,
			n('button.diffToggle', type='button', style='margin-bottom: 1em;', onclick='toggleTableDifference(this)', **{'data-diff': '0'}, c='Show Difference From Base') if not isBase else '',
			n('div.resultsPanel', [
				n('div.tableContainer', RawHtml(tableHtml))
			])
		])
		
		return str(mainBody)

	@cherrypy.expose
	def outputMap(self, scenarioId, month = None, stage = None):
		with serverDb() as db:
			scenario = db.queryRow('select projectId from scenario where id = ?', [scenarioId])
			baseScenarioId = db.queryValue('select id from scenario where projectId = ? and isBase', [scenario['projectId']])
		
		isBase = True if scenarioId == baseScenarioId else False
		
		monthsShort = 'Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec'.split(' ')
		allMonths = [monthsShort[i%len(monthsShort)] for i in range(24)]

		dateControl = n('div.dateSlider', style='--months: {};'.format(len(allMonths)), c=[
			n('input.month', type='range', min=0, max=len(allMonths)-1, step=1, value=month),
			n('datalist.ticks',
				[n('option', value=i, c=label, Class='selected' if month == i else None) for i,label in enumerate(allMonths)],
			),
		])
		
		mapControls = n('div.mapControls', [
			n('button', type='button', onclick='zoomMap({reset: true})', c='Reset View'),
			n('button', type='button', onclick='toggleBorders()', c='Toggle Borders'),
			n('button.diffToggle', type='button', onclick='toggleMapDifference(this)', c='Show Difference From Base') if not isBase else None,
		])
		
		bottomMapControls = n('div.mapControls.bottom', [dateControl])

		body = n('style', """
			.nzMap, .nzOutline { width: 500px; position: absolute; z-index: 5; }
			.nzOutline { z-index: 7; pointer-events: none; }
			.featureInfo { position: absolute; z-index: 10; background: white; border: solid 1px #ccc;
				padding: 10px; /*width: 400px; height: 300px;*/ }
			.featureInfo canvas { width: 400px; height: 300px; }
			""").str() 
		
		body += n('style.bordersOn', c="""
			/*.nzOutline polygon { stroke: none !important; }*/
			""").str() 
		
		body += n('style.bordersOff', c="""
			.nzMap polygon { stroke: none !important; }
			""").str()
		
		with open('_/images/10kmHexClippedNZTM.svg','r') as svgFile:
			body += svgFile.read()
		
		with open('_/images/nzoutline.svg','r') as svgFile:
			body += svgFile.read()

		body += str(n('script', """
			var mapData = {};
			var minVal = 0
			var maxVal = 0;
			var mapDiff = false;
			var scenarioId = """+json.dumps(scenarioId)+""";
			var baseScenarioId = """+json.dumps(baseScenarioId)+""";
			var stage = """+json.dumps(stage)+""";
			
			function formatRecord(rec) {
				let sigFigs = 4;
				var str = `
					<h2>${toHtml(rec.EA_Name)}</h2>
					<div><strong>mean/sq. km:</strong> <span>${toHtml(sigFig(rec.uPeSqKm, sigFigs))}</span></div>
					<div><br>Double click for dynamic view</div>
				`;
				return str;
			}
			
			function toggleMapDifference(button) {
				mapDiff = !mapDiff;
				if (mapDiff) {
					$(button).text('Hide Difference from Base');
				}
				else {
					$(button).text('Show Difference from Base');
				}
				updateMap();
			}
			
			
			$(document).ready(function() {
				$(document).on('mouseover', 'polygon', function(event) {
					// Always remove any existing popup first
					$('div.featureInfo').remove();

					$(this).css({'stroke': 'red'});

					let resultsPanel = $('.resultsPanel')[0].getBoundingClientRect();
					let code = $(this).attr('data-code');
					let details = mapData[code];

					// Exit if there's no data
					if (!details) return;

					// Ensure mapViewParent exists
					let $parent = $('.mapViewParent');
					if (!$parent.length) return;

					// Create and append the new popup
					let $fi = $('<div class="featureInfo">')
						.css({
							top: (event.pageY - resultsPanel.top) + 10,
							left: (event.pageX - resultsPanel.left) + 10,
							position: 'absolute' // just in case it's missing from your CSS
						})
						.html(formatRecord(details))
						.appendTo($parent)
						.show();
				})

				.on('mouseout', 'polygon', function(event) {
					$(this).css({'stroke': ''});
					$('div.featureInfo').remove(); // Clean up popup on mouseout
				})

				.on('dblclick', 'polygon', function(event) {
					let params = getCurrentParameters();
					window.location.href = changeQsUrl(window.location.href, {
						...params,
						type: 'point',
						locationCode: $(this).attr('data-code')
					});
				});

				
				$('input[type=range]').on('input', function(event) {
					updateMap();
				});
				
				updateMap();
			});
			"""))
			

		mainBody = n('div.resultsPanel.mapPanel', [
			mapControls,
			n('div.mapViewParent',
				n('div.mapView', RawHtml(body))
			),
			bottomMapControls,
		])

		return str(mainBody)
	
	@cherrypy.expose
	def outputTimeLine(self, scenarioId, locationCode, stage):
		with serverDb() as db:
			scenario = db.queryRow('select isBase, projectId from scenario where id = ?', [scenarioId])
			isBase = scenario['isBase']
			baseScenarioId = db.queryValue('select id from scenario where projectId = ? and isBase', [scenario['projectId']])
		
		locationCode = int(locationCode)
		places = self.getPlaces()
		
		timeData = self.getTimeData(scenarioId, stage, locationCode)
		baseTimeData = self.getTimeData(baseScenarioId, stage, locationCode)
		
		body = n('style', """
			.pointGraph { position: relative; height: 100%; }
			.pointGraph canvas { height: 100%; }
			.pointGraph .sampleHistogram { height: 100% !important; position: relative;  }
			.floatingGraph { width: 300px; height: 200px; }
			html, body { overflow: hidden; }
			.pathwayPointsSel, .itemsSel { display: none; }
			""").str() 
		
		body += str(n('script', '''
			var timeData = '''+timeData+''';
			var baseTimeData = '''+baseTimeData+''';
			let numMonths = timeData.length;
			
			$('button.diffGraph').on('click', function() {
				let baseDataset = timelineChart.data.datasets[1];
				if (baseDataset.hidden) {
					baseDataset.hidden = false;
					$(this).text('Hide Base');
				}
				else {
					baseDataset.hidden = true;
					$(this).text('Show Base');
				}
				timelineChart.update();
			});
							
			$(document).ready(function() {
			
				$('.location input[list]').on('change', function() {
					updateTimeData();
				});
				
				makeTimeline();
			});
				
			'''))
			
		mainBody = n('div',
			n('button.diffGraph', type='button', style='margin-bottom: 1em;', c='Show Base') if not isBase else None,
			n('div.graph', 
	 			n('div.pointGraph', RawHtml(body)),	
				n('div.location', n('h2', 'Location:', places[locationCode])),
			)
		)
		
		return str(mainBody)


	@cherrypy.expose
	def scenarioOutput(self, scenarioId, type = 'map', locationCode = None, month = 0, stage = 'accumulated'):
		with serverDb() as db:
			scenario = db.queryRow('select * from scenario where id = ?', [scenarioId])
			project = db.queryRow('select * from project where id = ?', [scenario['projectId']])
			
			if type == 'table':
				output = self.outputTable(scenarioId)
			elif type == 'map':
				output = self.outputMap(scenarioId, month = month, stage = stage)
			else:
				output = self.outputTimeLine(scenarioId, locationCode = locationCode, stage = stage)

			head = [
				n('script', '''
				var displayField = 'uPeSqKm';
				var projectId = '''+json.dumps(scenario['projectId'])+''';
				var scenarioId = '''+json.dumps(scenarioId)+''';
				var outputType = '''+json.dumps(type)+''';
							
				$(document).ready(function() {
					$('[name=displayValue]').on('click', function(event) {
						if (["Entry"].includes($(this).parent().text())) {
							window.location.href = '/scenarioOutput?scenarioId='+scenarioId+'&stage=entry&locationCode=null&type=table';
						}
						else if (outputType=='table' & !["Entry"].includes($(this).parent().text())) {
							window.location.href = changeQsUrl(window.location.href, {stage: $(this)[0].value, type: 'map'});
						}
						else {
							updateMap()
						}
					});
				});
				'''),
			]

			body = [
				n('h2', project['name']),
				n('h2', dataFieldEdit='/scenarioNameUpd?id={}&name='.format(id), c=n('span.value',toHtml(scenario['name']))),
				n('div.controls', []),
				n('div.outputPanels', c=[
					n('div.sideBar', c=[
						n('h2', 'Run Information'),
						n('section',
							n('div', 'Values to display:'),
							n('ul.displayValues',
								n('li', n('div', n('input', type='radio', name='displayValue', value='entry', checked=stage=='entry' or None), 'Entry')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='cs', checked=stage=='cs' or None), 'Climate Suitability')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='ls', checked=stage=='ls' or None), 'Land Suitability')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='habitat', checked=stage=='habitat' or None), 'Habitat Suitability')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='exposure', checked=stage=='exposure' or None), 'Exposure')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='disperse', checked=stage=='disperse' or None), 'Disperse')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='accumulated', checked=stage=='accumulated' or None), 'Accumulated')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='establishment', checked=stage=='establishment' or None), 'Establishments')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='spread', checked=stage=='spread' or None), 'Spread')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='consequencesEconomic', checked=stage=='consequences' or None), 'Consequences - Economic')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='consequencesEnvironmental', checked=stage=='consequences' or None), 'Consequences - Environment')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='consequencesHealth', checked=stage=='consequences' or None), 'Consequences - Health')),
								n('li', n('div', n('input', type='radio', name='displayValue', value='consequencesSocial', checked=stage=='consequences' or None), 'Consequences - Social')),
							),
						),
					]),
					n('div.main.scrollable', c=RawHtml(output)),
				]),
			]
			
			return runFileTemplate('basic.html', head = head, body = body, trail = [
				n('li',n('a', href='/project?id={}'.format(scenario['projectId']), c=[n('i', Class="fas fa-fw fa-bug"), 'Project'])),
				n('li', n('i', Class="fas fa-fw fa-chart-bar"), 'Output'),
			])
			



# This is so we can force anything returned into either a string (including nodes) or a json object (if dict/list/bool)
from cherrypy._cpdispatch import Dispatcher, LateParamPageHandler
class ForceStringDispatcher(Dispatcher):
	def __call__(self, path_info):
		"""Set handler and config for the current request."""
		request = cherrypy.serving.request
		func, vpath = self.find_handler(path_info)
		
		def func2(*args, **kwargs):
			val = func(*args, **kwargs)
			if isinstance(val, dict) or isinstance(val, list) or isinstance(val, bool):
				return json.dumps(val)
			elif not isinstance(val, str):
				return bytes(str(val), 'utf-8')
			return val

		if func:
			# Decode any leftover %2F in the virtual_path atoms.
			vpath = [x.replace('%2F', '/') for x in vpath]
			request.handler = LateParamPageHandler(func2, *vpath)
		else:
			request.handler = cherrypy.NotFound()

class IgnorePings(logging.Filter):
	def filter(self, record):
		return 'GET /getProcessList' not in record.getMessage()
			
def secureheaders():
	headers = cherrypy.response.headers
	headers['X-Frame-Options'] = 'SAMEORIGIN'
	headers['X-XSS-Protection'] = '1; mode=block'
	headers['Content-Security-Policy'] = "default-src='self'"
	
cherrypy.tools.secureheaders = cherrypy.Tool('before_finalize', secureheaders, priority=60)

if __name__=="__main__":
	cherrypy.config.update({
		'engine.autoreload.on': True,
		'server.socket_host': '0.0.0.0',
		'server.socket_port': settings['port']
		# 'log.screen': True,
		# 'log.access_file': '',
		# 'log.error_file': '',
		# 'server.daemonize': False
	})
	
	print()
	print("Running NZ Generic Risk Assessment Server 0.1")
	print(f"http://localhost:{settings['port']}/")
	print()
	print("Hit Ctrl-Break to stop the server")
	print()

	def open_page():
		if not os.path.exists('development_version.txt'):
			webbrowser.open_new_tab(f"http://localhost:{settings['port']}/")
			
	# cherrypy.engine.subscribe('start', open_page)
	cherrypy.quickstart(IBRAMServer(), '/', config={
		'/': {
			'tools.staticdir.root': os.path.abspath(os.getcwd()),
			'tools.staticdir.content_types': {'xdsl': 'application/octet-stream'},
			'tools.secureheaders.on': True,
		},
		'/_': {
			'tools.staticdir.on': True,
			'tools.staticdir.dir': './_'
		},
		# '/bns': {
			# 'tools.staticdir.on': True,
			# 'tools.staticdir.dir': './bns'
		# },
		'/outputs': {
			'tools.staticdir.on': True,
			'tools.staticdir.dir': settings['outputsDir']
		},
	})