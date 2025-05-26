"""

Interface layer to the database on the Java side.

"""

import sqlite3, re, math, sys

from openpyxl import load_workbook

class DB:
	def __init__(self, path):
		if hasattr(path,'execute'):
			self.conn = path
			self.conn.row_factory = sqlite3.Row
		else:
			self.conn = sqlite3.connect(path, isolation_level=None)
			self.conn.execute('pragma journal_mode=wal')
			self.conn.row_factory = sqlite3.Row
		self.debug = False
		
		self.addMoreFunctions()
	
	def __enter__(self):
		return self
	
	def __exit__(self, type, value, tb):
		self.conn.close()
		
	def addMoreFunctions(self):
		self.conn.create_function('_ln', 1, math.log)
		self.conn.create_function('_exp', 1, math.exp)
		self.conn.create_function('_pow', 2, math.pow)
	
		class Prod:
			def __init__(self):
				self.total = 1
			def step(self, value):
				self.total *= value
			def finalize(self):
				return self.total
		self.conn.create_aggregate('_product', 1, Prod)
	
	def escapeIdentifier(self, id):
		if id.find("\x00") >= 0:
			raise "NUL characters not allowed in SQL identifiers"
		
		return "\"" + id.replace("\"", "\"\"") + "\""
	
	def query(self, sql, params = None, ignoreErrors = False, keyType = "name"):
		if keyType == "index":
			self.conn.row_factory = None
		else:
			self.conn.row_factory = sqlite3.Row
		
		if params is None: params = []
		if self.debug:
			print("query:",sql,"<br>")
			print("params:",params,"<br>")
			
		if ignoreErrors:
			rs = None
			try:
				rs = self.conn.execute(sql, params)
			except: pass
		else:
			rs = self.conn.execute(sql, params)
		
		# Always commit
		self.conn.commit()
		
		return rs
	
	def handleValType(self, val):
		return bautils.convertNumber(val)
		'''try:
			val = int(val)
		except:
			try:
				val = float(val)
			except: pass
		
		return val'''
	
	# This will be efficient, simple and protect against injections
	def intParams(self, lst):
		sqlFrag = "("
		sep = ""
		for item in lst:
			int(item) # If not int, throw error
			sqlFrag += sep + str(item)
			sep = ","
		
		sqlFrag += ")"
		
		return sqlFrag
	
	# Creates the placeholders string. More general than above function, but involves
	# duplication and is less efficient
	def placeholders(self, lst):
		return "("+("?,"*len(lst))[:-1]+")"
	
	def setPlaceholders(self, fields, set = True):
		if not fields:  return ''
		
		return ('set ' if set else '')+(', '.join(f+' = ? ' for f in fields))

	def queryValue(self, sql, params = None, typeHandling = True, ignoreErrors = False):
		rs = self.query(sql, params, ignoreErrors)
		if rs is None: return None
		row = rs.fetchone()
		if row is None:  return None

		return row[0]
	
	# keyType can be "name" or "index"
	def queryRow(self, sql, params = None, keyType = "name", typeHandling = True, ignoreErrors = False):
		rs = self.query(sql, params, ignoreErrors, keyType)
		row = rs.fetchone()
		if row is None:  return None
		
		if keyType == "name":
			return dict(row)
		elif keyType == "index":
			return list(row)
	
	# keyType can be "name" or "index"
	def queryRows(self, sql, params = None, keyType = "name", typeHandling = True, oneD = False, ignoreErrors = False):
		rs = self.query(sql, params, ignoreErrors, keyType)

		if oneD:
			data = []
			for row in rs:
				data.extend(row)
			return data
		
		rows = rs.fetchall()
		if rows is None:  return None
		
		if keyType == "name":
			return [dict(r) for r in rows]
		elif keyType == "index":
			return [list(r) for r in rows]
		elif keyType == "indexWithHeaders":
			return [[list(r) for r in rows],rows[0].keys()]
	
	# Returns a map of field1 to field2 (which are the first and second fields, respectively, by default)
	def queryMap(self, sql, params = None, ignoreErrors = False, field1 = 0, field2 = 1):
		rs = self.query(sql, params, ignoreErrors)

		map = {}
		for row in rs:
			map[row[field1]] = row[field2]
		
		return map

	def update(self, table, upds, condition, params, ignore = []):
		if not params: params = []

		updStr = ""
		sep = ""
		i = 0
		for field,val in upds.items():
			if field in ignore:  continue
			params.insert(i, val)
			updStr += sep + field + "= ?"
			sep = ", "
			i += 1
	
		if condition:
			condition = " where "+condition
		
		sql = "update "+table+" set "+updStr+condition
		
		if self.debug:
			print("query:",sql,"<br>")
			print("params:",params,"<br>")
		
		self.query(sql, params)
	
	def lastInsertId(self):
		return self.queryValue('select last_insert_rowid()')
	
	# pk can be string (name of primary key), or an array of strings (names) for a "composite" key
	# replace always returns the row id, even for composite keys
	def replace(self, table, upds, pk, valid = None):
		if isinstance(pk, list):
			pks = pk
		else:
			pks = [pk]
		for pk in pks:
			if pk not in upds:
				print("Primary key '{}' not in |upds|. Use None or -1 if inserting.", pk)
		pkWheres = ' and '.join('{} = ?'.format(pk) for pk in pks)
		recordExists = self.queryValue("select 1 from {} where {}".format(table, pkWheres), [upds[pk] for pk in pks])

		# Filter down to just those that are in valid (if valid specified)
		if valid:
			oldUpds = upds
			upds = {}
			for k in valid:
				if k in oldUpds:
					v = oldUpds[k]
					upds[k] = v
		
		if not recordExists:
			insertFields = [k for k in upds.keys() if k not in pks] if (len(pks)==1 and isinstance(upds[pk],int) and upds[pk]<0) else upds.keys()
			fieldStr = ", ".join(self.escapeIdentifier(f) for f in insertFields)
			placeholdersStr = ("?,"*len(insertFields))[:-1]
			params = [upds[f] for f in insertFields]
			sql = "insert into {} ({}) values ({})".format(self.escapeIdentifier(table), fieldStr, placeholdersStr)
			if self.debug: print(sql, params)
			self.query(sql, params)
			
			return self.queryValue("select last_insert_rowid()")
		else:
			updateFields = upds.keys()
			fieldSetStr = ", ".join(self.escapeIdentifier(k)+" = ?" for k in updateFields)
			params = [upds[f] for f in updateFields] + [upds[pk] for pk in pks]
			sql = "update {} set {} where {}".format(self.escapeIdentifier(table), fieldSetStr, pkWheres)
			if self.debug: print(sql, params)
			self.query(sql, params)
			
			if len(pks)==1:
				return upds[pk]
			else:
				return self.queryValue('select rowid from {} where {}'.format(self.escapeIdentifier(table), pkWheres), [upds[pk] for pk in pks])
	
	'''
	changeSet structure should be: array(
		"changed" => array(
			<pkValue> => array("field1" => val, etc.),
			<pkValue2> etc.
		),
		"inserted" => array(
			array("field1" => val, etc.),
			array("field1" => val, etc.),
			etc.
		),
		"deleted" => array(<pkValue>, <pkValue2>, etc.)
	)
	'''
	def applyTableChanges(self, table, changeSet, pkName):
		newPkMap = {}
		
		for pk, record in changeSet["changed"].items():
			record[pkName] = int(pk)
			self.replace(table, record, pkName)

		for record in changeSet["inserted"]:
			oldPk = record[pkName]
			record[pkName] = None
			newPk = self.replace(table, record, pkName)
			newPkMap[oldPk] = newPk

		for pk in changeSet["deleted"]:
			self.query("delete from "+table+" where "+pkName+" = ?", [pk])

		return newPkMap

	def tableExists(self, table):
		return (self.queryValue("select 1 from sqlite_master where type='table' and name = ?", [table]) is not None)
		
	def fieldNames(self, rs):
		return [r[0] for r in rs.description]
	
	# inTable is either a string name of a table, or a rs
	def writeCsv(self, inTable, outCsvFn, fields = "*", where = None):
		import csv
		with open(outCsvFn, "w", newline='') as outCsvFile:
			
			if type(inTable) == str:
				rs = self.query("select "+",".join(fields)+" from "+inTable+((" where "+where) if where else ""))
			else:
				rs = inTable
			
			# Write headers
			if fields == "*":
				headers = self.fieldNames(rs)
			else:
				headers = fields
			outCsv = csv.DictWriter(outCsvFile, headers)
			outCsv.writeheader()
			
			for row in rs:
				row = dict(row)
				outCsv.writerow(dict((k,row[k]) for k in headers if k in row))

			rs.close()


	def coerceValue(self, val):
		try:
			val = float(val)
			val = "{:g}".format(val)
		except: pass

		try: val = int(val)
		except:
			try: val = float(val)
			except: pass
		
		return val
	
	def _readExcelRow(self, sheet, rowI, headers = None):
		if headers:
			d = {}
			for c in range(sheet.max_column):
				cell = sheet.cell(row=rowI+1, column=c+1)
				d[headers[c]] = self.coerceValue(cell.value)
			return d
		else:
			l = []
			for c in range(sheet.max_column):
				cell = sheet.cell(row=rowI+1, column=c+1)
				l.append(self.coerceValue(cell.value))
			return l
	
	# def loadDataFromExcel(self, wb):
	# 	errors = []
		
	# 	for s in wb:
	# 		# Replace all non-word characters with underscore
	# 		sheetName = re.sub(r'\W', '_', s.title)
			
	# 		print("Checking '{}'".format(sheetName))
	# 		if self.tableExists(sheetName):
	# 			for c in range(s.max_column):
	# 				# Replace all non-word characters with underscore
	# 				colName = re.sub(r'\W', '_', s.cell(row=1,column=c+1).value)
	# 				# print(colName)
					
	# 				# Skip hidden columns
	# 				if re.search(r'^hidden_', colName):  continue
					
	# 				hasColumn = False
	# 				try:
	# 					self.query("select {} from {} limit 1".format(colName, sheetName))
	# 					hasColumn = True
	# 				except: pass

	# 				if not hasColumn:
	# 					errors.append("'{}' table does not have the '{}' column".format(sheetName,colName))
	# 		else:
	# 			errors.append("'{}' table does not exist".format(sheetName))
		
	# 	if errors:
	# 		print("Aborting due to following errors:")
	# 		print("\n".join(errors))
	# 		sys.exit(0)
	# 	else:
	# 		print("Schema checks passed. Proceeding with data loading.")
		
	# 	# If OK, proceed with loading data
	# 	for s in wb:
			
	# 		headers = None
	# 		for r in range(s.max_row):
	# 			if r==0:
	# 				headers = self._readExcelRow(s, r)
	# 			else:
	# 				row = self._readExcelRow(s, r, headers)
	# 				# Filter out hidden columns
	# 				row = {k:v for k,v in row.items() if not re.search(r'^hidden_', k)}
	# 				# Exit loop at the first blank row
	# 				if all(v is None for k,v in row.items()): break 
	# 				self.replace(s.title, row, headers[0])
					


	def loadDataFromExcel(self, wb):
		errors = []

		for s in wb:
			sheetName = re.sub(r'\W', '_', s.title)
			print("Checking '{}'".format(sheetName))

			if self.tableExists(sheetName):
				for c in range(s.max_column):
					second_cell = s.cell(row=1, column=c+1).value
					if second_cell is None:  # Skip column with empty first cell
						continue

					colName = re.sub(r'\W', '_', second_cell)

					if re.search(r'^hidden_', colName):
						continue

					hasColumn = False
					try:
						self.query("select {} from {} limit 1".format(colName, sheetName))
						hasColumn = True
					except:
						pass

					if not hasColumn:
						errors.append("'{}' table does not have the '{}' column".format(sheetName, colName))
			else:
				errors.append("'{}' table does not exist".format(sheetName))

		if errors:
			print("Aborting due to following errors:")
			print("\n".join(errors))
			sys.exit(0)
		else:
			print("Schema checks passed. Proceeding with data loading.")

		# If OK, proceed with loading data
		for s in wb:
			headers = None
			for r in range(s.max_row):

				if r == 0:
					headers = self._readExcelRow(s, r)
				else:
					row = self._readExcelRow(s, r, headers)
					row = {k: v for k, v in row.items() if not re.search(r'^hidden_', k)}
					
					if all(v is None for k, v in row.items()):
						break
					
					self.replace(s.title, row, headers[0])


